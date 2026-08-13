#!/usr/bin/env python3
"""
MAX Weekly Datasheet — Digital-Footprint Audit
================================================
Standalone post-processor for a generated Weekly_Vendor_Report datasheet.xlsx.
Runs the SAME Driftnet ownership check as max-hostname-triage / max-findings-triage
against every domain in the per-row finding sheets (Critical Indicators, High
Indicators, New Findings (7d)) and either:

  --mode annotate  Add a "Digital Footprint Verdicts" summary sheet; leave the
                    original sheets untouched (safest — use when you just want
                    visibility before deciding what to do).
  --mode clean     Also move SHARED_INFRA_MISMATCH / NO_DNS_RECORD rows out of
                    the original sheets into a new "Excluded — Unverified
                    Footprint" sheet (never deleted outright — always auditable),
                    and tag STALE/AMBIGUOUS rows in-place as held rather than
                    silently leaving them looking identical to a confirmed row.
  --mode both      Do both. This is the default — the whole point of this skill
                    is that a customer should never see a footprint claim this
                    datasheet can't back up, whether via a cleaner primary sheet
                    or a doubly-visible verdict.

Why this exists as its own skill, separate from max-hostname-triage /
max-findings-triage: those two operate on LIVE MAX findings via the partner API,
gating what gets marked report=true before a finding is even eligible to appear
in a generated report. But weekly report generation still runs independently and
a mis-attributed finding can and does end up in a datasheet regardless (a
`weekly-report-generator` run isn't gated by workstation triage state) — so
reports need their own last-mile check, not just an assumption that upstream
triage caught everything. Run this on ANY datasheet before it goes to a customer,
even one you didn't personally triage.

Designed to slot into the standard pipeline with a one-line addition — see the
"Pipeline integration" section in SKILL.md — not as a parallel, separately-run
process analysts have to remember.

Usage:
    python3 audit_datasheet.py --in datasheet.xlsx --out datasheet_audited.xlsx
    python3 audit_datasheet.py --in datasheet.xlsx --out datasheet_audited.xlsx --mode annotate
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill
import requests

DRIFTNET_BASE_URL = "https://api.driftnet.io/v1"   # confirm paths against
                                                     # docs/DRIFTNET_API_REFERENCE.md

SHARED_INFRA_MARKERS = [
    "akamai", "cloudflare", "fastly", "amazon", "aws", "microsoft azure",
    "google cloud", "digitalocean", "ovh", "securityscorecard",
]

# Sheets this audit understands: one row per (domain, issue) — the sheets that
# actually surface as customer-facing finding tables in the generated deck.
PER_ROW_SHEETS = {
    "Critical Indicators": {"domain_col": "DOMAIN", "vendor_col": None},
    "High Indicators":     {"domain_col": "DOMAIN", "vendor_col": None},
    "New Findings (7d)":   {"domain_col": "DOMAIN", "vendor_col": "VENDOR"},
}

EXCLUDE_VERDICTS = {"SHARED_INFRA_MISMATCH", "NO_DNS_RECORD", "HONEYPOT_INFRASTRUCTURE", "VOLATILE_SHARED_PLATFORM"}
HOLD_VERDICTS    = {"STALE_NO_SCAN", "UNVERIFIABLE", "AMBIGUOUS"}

HOLD_FILL     = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
EXCLUDE_FONT  = Font(color="B3311F")
CONFIRM_FONT  = Font(color="167A56")


def root_domain(d):
    parts = str(d).lower().strip().rstrip(".").split(".")
    return parts[-2] if len(parts) >= 2 else parts[0]


def _driftnet_get(token, path, params):
    if not token:
        return None
    try:
        r = requests.get(f"{DRIFTNET_BASE_URL}{path}",
                          headers={"Authorization": f"Bearer {token}"},
                          params=params, timeout=15)
        return r.json() if r.status_code == 200 else None
    except requests.RequestException:
        return None


def driftnet_ownership_verdict(token, hostname, claimed_owner):
    """Same verdict engine as max-hostname-triage / max-findings-triage. Returns
    (verdict, reason, evidence_dict)."""
    owner = root_domain(claimed_owner)

    dns = _driftnet_get(token, "/dns/forward", {"expression": f"host={hostname}"})
    if dns is None:
        return "UNVERIFIABLE", "Driftnet DNS lookup failed or token missing.", {}

    records = dns.get("reports") or dns.get("records") or []
    if not records:
        return "NO_DNS_RECORD", f"'{hostname}' has no DNS presence in Driftnet's corpus.", {}

    def _a_ips(rec):
        return {i["value"] for i in rec.get("items", [])
                if i.get("type") == "ip" and i.get("context", "").startswith("dns-a")}

    # Check A-record stability across whatever snapshots came back (newest first).
    # A domain whose resolved IPs churn completely between snapshots fronts a
    # rotating/anycast service, not a fixed host — pinning a single-asset
    # severity score to "the IP behind it" isn't reproducible. Only meaningful
    # with 2+ snapshots; with just one, fall through to the single-snapshot check.
    snap_ip_sets = [s for s in (_a_ips(r) for r in records) if s]
    if len(snap_ip_sets) >= 2 and not all(s == snap_ip_sets[0] for s in snap_ip_sets[1:]):
        return "VOLATILE_SHARED_PLATFORM", (
            f"Resolved IPs changed across all {len(snap_ip_sets)} available snapshots — "
            "domain fronts a rotating/anycast service, not a fixed host."
        ), {}

    ips = list(snap_ip_sets[0]) if snap_ip_sets else []
    if not ips:
        return "AMBIGUOUS", "DNS record found but no A records parsed.", {}

    scan = _driftnet_get(token, "/scans/protocols", {"ip": ips[0], "most_recent": "true"})
    scan_records = (scan or {}).get("reports") or (scan or {}).get("records") or []
    if not scan_records:
        return "UNVERIFIABLE", f"No recent passive scan of {ips[0]}.", {}

    evidence = {}
    hay_parts = []
    is_honeypot = False
    for rep in scan_records:
        for item in rep.get("items", []):
            t, v = item.get("type"), item.get("value")
            if t == "tag" and item.get("context") == "driftnet" and v == "honeypot":
                is_honeypot = True
            if t in ("entity", "asn", "subject", "host") and v:
                hay_parts.append(str(v))
                evidence.setdefault(t, str(v))
    hay = " ".join(hay_parts).lower()

    # Checked before ownership matching: a honeypot can legitimately carry a real
    # cert for the claimed vendor (WAF/deception nodes are sometimes provisioned
    # with the real domain's cert precisely to look authentic) — that's not
    # "confirmed ownership", it's the opposite. A finding observed against a
    # honeypot describes deliberately-planted bait, not the vendor's real exposure.
    if is_honeypot:
        return "HONEYPOT_INFRASTRUCTURE", f"{ips[0]} is tagged by Driftnet as deception/honeypot infrastructure — any issue signature observed here does not reflect '{claimed_owner}'s real exposure.", evidence

    owner_present = owner in hay
    shared_hit = next((m for m in SHARED_INFRA_MARKERS if m in hay and m not in owner), None)

    if shared_hit and not owner_present:
        return "SHARED_INFRA_MISMATCH", f"Latest scan of {ips[0]} attributes it to '{shared_hit}', not '{claimed_owner}'.", evidence
    if owner_present:
        return "CONFIRMED", f"Latest scan of {ips[0]} attribution matches '{claimed_owner}'.", evidence
    return "AMBIGUOUS", "No shared-infra marker and no owner-name match.", evidence


def read_sheet_rows(ws):
    rows = list(ws.iter_rows())
    if not rows:
        return None, []
    header_cells = rows[0]
    headers = [c.value for c in header_cells]
    return headers, rows[1:]


def run_audit(in_path, out_path, mode, driftnet_token):
    wb = openpyxl.load_workbook(in_path, data_only=True)

    # ── Pass 1: collect every unique (domain, claimed_owner) across understood sheets
    unique_pairs = {}   # domain -> claimed_owner (first vendor name seen, else domain itself)
    per_sheet_rows = {}  # sheet_name -> (headers, [row cells])

    for sheet_name, cfg in PER_ROW_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers, rows = read_sheet_rows(ws)
        if not headers:
            continue
        per_sheet_rows[sheet_name] = (headers, rows)
        domain_idx = headers.index(cfg["domain_col"]) if cfg["domain_col"] in headers else None
        vendor_idx = headers.index(cfg["vendor_col"]) if cfg["vendor_col"] and cfg["vendor_col"] in headers else None
        if domain_idx is None:
            continue
        for row in rows:
            domain = row[domain_idx].value
            if not domain or str(domain).strip() in ("", "No indicators to Report", "No breaches to report"):
                continue
            vendor = row[vendor_idx].value if vendor_idx is not None else None
            unique_pairs.setdefault(str(domain).strip(), str(vendor).strip() if vendor else str(domain).strip())

    print(f"Unique domains found across {len(per_sheet_rows)} finding sheet(s): {len(unique_pairs)}")
    if not driftnet_token:
        print("⚠  No Driftnet token — every domain will come back UNVERIFIABLE and be held, not cleared.")

    # ── Pass 2: verdict per unique domain (deduped Driftnet calls) ──────────────
    verdicts = {}
    for i, (domain, owner) in enumerate(unique_pairs.items(), 1):
        v, reason, evidence = driftnet_ownership_verdict(driftnet_token, domain, owner)
        verdicts[domain] = {"verdict": v, "reason": reason, "evidence": evidence, "claimed_owner": owner}
        print(f"  [{i}/{len(unique_pairs)}] {domain:<55} {v}")

    counts = {}
    for v in verdicts.values():
        counts[v["verdict"]] = counts.get(v["verdict"], 0) + 1
    print("\nVerdict distribution:", counts)

    # ── Build "Digital Footprint Verdicts" summary sheet (always, both modes) ──
    if "Digital Footprint Verdicts" in wb.sheetnames:
        del wb["Digital Footprint Verdicts"]
    vs = wb.create_sheet("Digital Footprint Verdicts")
    vs.append(["DOMAIN", "CLAIMED OWNER", "VERDICT", "ENTITY", "ASN", "REASON", "RECOMMENDED ACTION"])
    action_map = {
        "CONFIRMED": "None — ownership verified",
        "SHARED_INFRA_MISMATCH": "Remove from customer-facing report or re-attribute",
        "NO_DNS_RECORD": "Remove — not a verifiable internet asset",
        "HONEYPOT_INFRASTRUCTURE": "Remove — signature observed against deception infrastructure, not the vendor",
        "VOLATILE_SHARED_PLATFORM": "Remove — domain fronts a rotating/anycast service, not a fixed asset",
        "STALE_NO_SCAN": "Hold — re-check next cycle, do not escalate",
        "UNVERIFIABLE": "Hold — Driftnet call failed or token missing",
        "AMBIGUOUS": "Hold — needs analyst judgment",
    }
    for domain, v in sorted(verdicts.items(), key=lambda kv: kv[1]["verdict"]):
        vs.append([domain, v["claimed_owner"], v["verdict"],
                   v["evidence"].get("entity", ""), v["evidence"].get("asn", ""),
                   v["reason"], action_map.get(v["verdict"], "")])
    for cell in vs[1]:
        cell.font = Font(bold=True)

    # ── mode=clean: move/tag rows in the original per-row sheets ────────────────
    if mode in ("clean", "both"):
        if "Excluded — Unverified Footprint" in wb.sheetnames:
            del wb["Excluded — Unverified Footprint"]
        excl_ws = wb.create_sheet("Excluded — Unverified Footprint")
        excl_ws.append(["SOURCE SHEET", "DOMAIN", "VERDICT", "REASON", "ORIGINAL ROW"])
        excl_ws[1][0].font = Font(bold=True)

        for sheet_name, cfg in PER_ROW_SHEETS.items():
            if sheet_name not in per_sheet_rows:
                continue
            headers, rows = per_sheet_rows[sheet_name]
            domain_idx = headers.index(cfg["domain_col"])
            ws = wb[sheet_name]

            keep_row_nums = []
            for row in rows:
                domain = row[domain_idx].value
                v = verdicts.get(str(domain).strip(), {}).get("verdict") if domain else None
                if v in EXCLUDE_VERDICTS:
                    excl_ws.append([sheet_name, domain, v, verdicts[str(domain).strip()]["reason"],
                                     str([c.value for c in row])])
                    for cell in row:
                        cell.font = EXCLUDE_FONT
                elif v in HOLD_VERDICTS:
                    for cell in row:
                        cell.fill = HOLD_FILL
                    keep_row_nums.append(row[0].row)
                else:
                    keep_row_nums.append(row[0].row)

            # Delete excluded rows bottom-up so row indices don't shift under us
            all_row_nums = [r[0].row for r in rows]
            delete_nums = sorted(set(all_row_nums) - set(keep_row_nums), reverse=True)
            for rn in delete_nums:
                ws.delete_rows(rn, 1)

        print(f"\nExcluded rows moved to 'Excluded — Unverified Footprint': "
              f"{excl_ws.max_row - 1}")

    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    return verdicts


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--in", dest="in_path", required=True)
    parser.add_argument("--out", dest="out_path", required=True)
    parser.add_argument("--mode", choices=["annotate", "clean", "both"], default="both")
    args = parser.parse_args()

    token = os.environ.get("DRIFTNET_API_TOKEN")
    run_audit(args.in_path, args.out_path, args.mode, token)
